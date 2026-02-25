import os
import sys
import re
import logging
import hashlib
import tempfile
from io import BytesIO
from pathlib import Path
from urllib.parse import urlparse
from concurrent.futures import ThreadPoolExecutor, as_completed

import numpy as np
import requests
import psycopg2
from PIL import Image, ImageEnhance, ImageFilter, ImageOps
from tqdm import tqdm

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s"
)
logger = logging.getLogger(__name__)

SUPPORTED_FORMATS = {".png", ".jpg", ".jpeg", ".bmp", ".tiff", ".tif", ".webp"}
CACHE_DIR = Path(__file__).parent / ".img_cache"
TEMP_DIR = Path(__file__).parent / ".tmp_preprocess"


class DBConfig:
    HOST = os.getenv("DB_HOST", "localhost")
    PORT = int(os.getenv("DB_PORT", "5432"))
    NAME = os.getenv("DB_NAME", "docdb")
    USER = os.getenv("DB_USER", "admin")
    PASSWORD = os.getenv("DB_PASSWORD", "admin123")

    @classmethod
    def dsn(cls):
        return {
            "host": cls.HOST,
            "port": cls.PORT,
            "dbname": cls.NAME,
            "user": cls.USER,
            "password": cls.PASSWORD,
        }


class DocumentRepository:
    def __init__(self, conn):
        self._conn = conn

    def fetch_unconverted(self):
        with self._conn.cursor() as cur:
            cur.execute(
                "SELECT id, images FROM document "
                "WHERE latex IS NULL OR latex = ''"
            )
            return cur.fetchall()

    def update_latex(self, doc_id: int, latex: str):
        with self._conn.cursor() as cur:
            cur.execute(
                "UPDATE document SET latex = %s WHERE id = %s",
                (latex, doc_id),
            )
        self._conn.commit()

    def insert_image(self, image_path: str):
        with self._conn.cursor() as cur:
            cur.execute(
                "INSERT INTO document (images) VALUES (%s) RETURNING id",
                (image_path,),
            )
            doc_id = cur.fetchone()[0]
        self._conn.commit()
        return doc_id


class ImageLoader:
    """Tải ảnh từ local path hoặc URL, cache URL đã download."""

    URL_PATTERN = re.compile(r'^https?://')
    DOWNLOAD_TIMEOUT = 30
    MAX_DOWNLOAD_WORKERS = 8

    @classmethod
    def is_url(cls, source: str) -> bool:
        return bool(cls.URL_PATTERN.match(source))

    @classmethod
    def _url_to_cache_path(cls, url: str) -> Path:
        url_hash = hashlib.md5(url.encode()).hexdigest()
        parsed = urlparse(url)
        ext = Path(parsed.path).suffix.lower()
        if ext not in SUPPORTED_FORMATS:
            ext = ".png"
        return CACHE_DIR / f"{url_hash}{ext}"

    @classmethod
    def download_single(cls, url: str) -> Path:
        cache_path = cls._url_to_cache_path(url)
        if cache_path.exists():
            return cache_path

        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        resp = requests.get(url, timeout=cls.DOWNLOAD_TIMEOUT, stream=True)
        resp.raise_for_status()

        content_type = resp.headers.get("Content-Type", "")
        if not content_type.startswith("image/"):
            img = Image.open(BytesIO(resp.content))
            img.save(cache_path)
        else:
            with open(cache_path, "wb") as f:
                for chunk in resp.iter_content(8192):
                    f.write(chunk)

        return cache_path

    @classmethod
    def download_batch(cls, urls: list[str]) -> dict[str, Path]:
        results = {}
        urls_to_download = []

        for url in urls:
            cache_path = cls._url_to_cache_path(url)
            if cache_path.exists():
                results[url] = cache_path
            else:
                urls_to_download.append(url)

        if not urls_to_download:
            return results

        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        workers = min(cls.MAX_DOWNLOAD_WORKERS, len(urls_to_download))

        with ThreadPoolExecutor(max_workers=workers) as pool:
            future_map = {
                pool.submit(cls.download_single, url): url
                for url in urls_to_download
            }
            for future in tqdm(
                as_completed(future_map),
                total=len(urls_to_download),
                desc="Download ảnh",
                unit="file",
            ):
                url = future_map[future]
                try:
                    results[url] = future.result()
                except Exception as e:
                    logger.error(f"[DOWNLOAD FAIL] {url} | {e}")

        return results

    @classmethod
    def resolve_path(cls, source: str) -> str:
        if cls.is_url(source):
            return str(cls.download_single(source))
        return source

    @classmethod
    def load_image(cls, source: str) -> Image.Image:
        path = cls.resolve_path(source)
        if not Path(path).exists():
            raise FileNotFoundError(f"Không tìm thấy ảnh: {source}")
        img = Image.open(path)
        if img.mode in ("RGBA", "P"):
            bg = Image.new("RGB", img.size, (255, 255, 255))
            if img.mode == "P":
                img = img.convert("RGBA")
            bg.paste(img, mask=img.split()[3] if "A" in img.getbands() else None)
            return bg
        return img.convert("RGB")


class ImagePreprocessor:

    @staticmethod
    def enhance(img: Image.Image) -> Image.Image:
        img = ImageOps.autocontrast(img, cutoff=1)
        img = ImageEnhance.Contrast(img).enhance(1.5)
        img = ImageEnhance.Sharpness(img).enhance(2.0)
        return img

    @staticmethod
    def denoise(img: Image.Image) -> Image.Image:
        return img.filter(ImageFilter.MedianFilter(size=3))

    @staticmethod
    def to_binary(img: Image.Image, threshold: int = 180) -> Image.Image:
        gray = img.convert("L")
        return gray.point(lambda x: 255 if x > threshold else 0, "L").convert("RGB")

    @staticmethod
    def optimal_resize(img: Image.Image, max_dim: int = 1024) -> Image.Image:
        w, h = img.size
        if max(w, h) > max_dim:
            ratio = max_dim / max(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        if min(w, h) < 32:
            ratio = 32 / min(w, h)
            img = img.resize((int(w * ratio), int(h * ratio)), Image.LANCZOS)
        return img

    @classmethod
    def standard(cls, img: Image.Image) -> Image.Image:
        img = cls.optimal_resize(img)
        img = cls.enhance(img)
        img = cls.denoise(img)
        return img

    @classmethod
    def save_preprocessed(cls, img: Image.Image, suffix: str = "") -> str:
        TEMP_DIR.mkdir(parents=True, exist_ok=True)
        tmp = TEMP_DIR / f"pre_{id(img)}_{suffix}.png"
        img.save(tmp)
        return str(tmp)


class LatexPostProcessor:
    """Dọn dẹp LaTeX output từ OCR."""

    ARTIFACTS = [
        (re.compile(r'^\\(scriptstyle|textstyle)\s*'), ''),
        (re.compile(r'\\;\s*\\;\s*\\;\s*\\;\s*\\;'), ' '),
        (re.compile(r'\s{2,}'), ' '),
    ]

    @classmethod
    def clean(cls, latex: str) -> str:
        if not latex:
            return latex
        result = latex.strip()
        for pattern, replacement in cls.ARTIFACTS:
            result = pattern.sub(replacement, result)
        return result.strip()


class TexTellerEngine:
    """TexTeller - train trên 80M cặp dữ liệu, chính xác hơn pix2tex."""

    def __init__(self):
        from texteller import load_model, load_tokenizer, img2latex as _img2latex
        logger.info("Đang tải model TexTeller...")
        self._model = load_model(use_onnx=False)
        self._tokenizer = load_tokenizer()
        self._img2latex = _img2latex
        logger.info("TexTeller đã sẵn sàng.")

    def predict(self, image_path: str) -> str:
        results = self._img2latex(self._model, self._tokenizer, [image_path])
        return results[0] if results else ""

    def predict_from_image(self, img: Image.Image) -> str:
        tmp_path = ImagePreprocessor.save_preprocessed(img, "texteller")
        try:
            return self.predict(tmp_path)
        finally:
            Path(tmp_path).unlink(missing_ok=True)


class Pix2TexEngine:
    """pix2tex (LaTeX-OCR) - fallback engine."""

    def __init__(self):
        from pix2tex.cli import LatexOCR
        logger.info("Đang tải model pix2tex (fallback)...")
        self._model = LatexOCR()
        self._model.args.temperature = 0.05
        logger.info("pix2tex đã sẵn sàng.")

    def predict(self, image_path: str) -> str:
        img = Image.open(image_path).convert("RGB")
        return self._model(img)

    def predict_from_image(self, img: Image.Image) -> str:
        return self._model(img)


class TextOCREngine:
    """EasyOCR - OCR văn bản tiếng Việt và tiếng Anh, giữ nguyên dấu."""

    VIETNAMESE_CHARS = re.compile(
        r'[àáạảãâầấậẩẫăằắặẳẵèéẹẻẽêềếệểễìíịỉĩ'
        r'òóọỏõôồốộổỗơờớợởỡ'
        r'ùúụủũưừứựửữỳýỵỷỹđ'
        r'ÀÁẠẢÃÂẦẤẬẨẪĂẰẮẶẲẴÈÉẸẺẼÊỀẾỆỂỄÌÍỊỈĨ'
        r'ÒÓỌỎÕÔỒỐỘỔỖƠỜỚỢỞỠ'
        r'ÙÚỤỦŨƯỪỨỰỬỮỲÝỴỶỸĐ]'
    )

    def __init__(self):
        import easyocr
        logger.info("Đang tải model EasyOCR (Việt + Anh)...")
        self._reader = easyocr.Reader(['vi', 'en'], gpu=False)
        logger.info("EasyOCR đã sẵn sàng.")

    def predict(self, image_path: str) -> str:
        results = self._reader.readtext(image_path, detail=0, paragraph=True)
        return "\n".join(results) if results else ""

    def predict_from_image(self, img: Image.Image) -> str:
        img_array = np.array(img)
        results = self._reader.readtext(img_array, detail=0, paragraph=True)
        return "\n".join(results) if results else ""

    @classmethod
    def has_vietnamese(cls, text: str) -> bool:
        return bool(cls.VIETNAMESE_CHARS.search(text))


class ImageToLatexConverter:
    """
    Hybrid OCR engine:
    1. EasyOCR      - văn bản tiếng Việt + tiếng Anh (giữ dấu)
    2. TexTeller    - công thức toán học → LaTeX
    3. pix2tex      - fallback công thức toán

    Chiến lược routing:
    - Phát hiện tiếng Việt → ưu tiên EasyOCR
    - Không có tiếng Việt  → thử TexTeller/pix2tex cho LaTeX
    - Tất cả fail          → fallback về EasyOCR text
    """

    def __init__(self):
        self._text_engine = TextOCREngine()
        self._math_engine = TexTellerEngine()
        self._fallback = None

    def _load_fallback(self):
        if self._fallback is None:
            self._fallback = Pix2TexEngine()
        return self._fallback

    def convert(self, source: str) -> str:
        local_path = ImageLoader.resolve_path(source)

        text_result = self._try_text_ocr(local_path, source)
        has_vietnamese = text_result and TextOCREngine.has_vietnamese(text_result)
        has_text = text_result and len(text_result.strip()) > 3

        if has_vietnamese:
            logger.info(f"  Phát hiện tiếng Việt → EasyOCR: {source}")
            return text_result

        math_result = self._try_math_ocr(local_path, source)
        if math_result:
            return LatexPostProcessor.clean(math_result)

        if has_text:
            return text_result

        raise RuntimeError(f"Không thể OCR ảnh: {source}")

    def _try_text_ocr(self, local_path: str, source: str) -> str | None:
        try:
            result = self._text_engine.predict(local_path)
            if result and result.strip():
                return result
        except Exception as e:
            logger.debug(f"  EasyOCR lỗi: {e}")

        try:
            raw_img = ImageLoader.load_image(source)
            preprocessed = ImagePreprocessor.standard(raw_img.copy())
            result = self._text_engine.predict_from_image(preprocessed)
            if result and result.strip():
                return result
        except Exception as e:
            logger.debug(f"  EasyOCR preprocessed lỗi: {e}")

        return None

    def _try_math_ocr(self, local_path: str, source: str) -> str | None:
        latex = self._try_engine(self._math_engine, local_path, source)
        if latex:
            return latex

        try:
            raw_img = ImageLoader.load_image(source)
            preprocessed = ImagePreprocessor.standard(raw_img.copy())
            latex = self._try_with_preprocessed(self._math_engine, preprocessed)
            if latex:
                return latex
        except Exception:
            pass

        logger.info(f"  TexTeller fail -> fallback pix2tex: {source}")
        fallback = self._load_fallback()
        latex = self._try_engine(fallback, local_path, source)
        if latex:
            return latex

        try:
            raw_img = ImageLoader.load_image(source)
            preprocessed = ImagePreprocessor.standard(raw_img.copy())
            latex = self._try_with_preprocessed(fallback, preprocessed)
            if latex:
                return latex
        except Exception:
            pass

        return None

    @staticmethod
    def _try_engine(engine, path: str, source: str) -> str | None:
        try:
            result = engine.predict(path)
            if result and result.strip():
                return result
        except Exception as e:
            logger.debug(f"  Engine {engine.__class__.__name__} lỗi: {e}")
        return None

    @staticmethod
    def _try_with_preprocessed(engine, img: Image.Image) -> str | None:
        try:
            result = engine.predict_from_image(img)
            if result and result.strip():
                return result
        except Exception as e:
            logger.debug(f"  Preprocessed fail: {e}")
        return None


class App:
    def __init__(self):
        self._conn = psycopg2.connect(**DBConfig.dsn())
        self._repo = DocumentRepository(self._conn)
        self._converter = ImageToLatexConverter()

    def insert_images(self, sources: list[str]):
        for s in sources:
            doc_id = self._repo.insert_image(s)
            tag = "URL" if ImageLoader.is_url(s) else "LOCAL"
            logger.info(f"[{tag}] Đã thêm: id={doc_id}, source={s}")

    def convert_all(self):
        rows = self._repo.fetch_unconverted()
        if not rows:
            logger.info("Không có ảnh nào cần convert.")
            return

        total = len(rows)
        logger.info(f"Tìm thấy {total} ảnh cần convert.")

        urls = [src for _, src in rows if ImageLoader.is_url(src)]
        if urls:
            logger.info(f"Đang download {len(urls)} ảnh từ URL...")
            ImageLoader.download_batch(urls)

        success = 0
        for doc_id, source in tqdm(rows, desc="Convert OCR", unit="ảnh"):
            try:
                latex = self._converter.convert(source)
                self._repo.update_latex(doc_id, latex)
                preview = latex[:80] + "..." if len(latex) > 80 else latex
                tqdm.write(f"  [OK] id={doc_id} | {source} -> {preview}")
                success += 1
            except Exception as e:
                tqdm.write(f"  [FAIL] id={doc_id} | {source} | {e}")

        logger.info(f"Kết quả: {success}/{total} ảnh convert thành công.")

    def close(self):
        self._conn.close()
        TEMP_DIR_PATH = Path(TEMP_DIR)
        if TEMP_DIR_PATH.exists():
            for f in TEMP_DIR_PATH.iterdir():
                f.unlink(missing_ok=True)
            TEMP_DIR_PATH.rmdir()


def export_to_excel(output_path: str = "document_database.xlsx"):
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    conn = psycopg2.connect(**DBConfig.dsn())
    cur = conn.cursor()
    cur.execute(
        "SELECT id, images, latex, created_at, updated_at "
        "FROM document ORDER BY id"
    )
    rows = cur.fetchall()
    cur.close()
    conn.close()

    if not rows:
        logger.info("Database trống, không có gì để export.")
        return

    wb = Workbook()
    ws = wb.active
    ws.title = "Documents"

    headers = ["ID", "Ảnh (đường dẫn)", "Kết quả OCR", "Ngày tạo", "Ngày cập nhật"]
    header_font = Font(name="Times New Roman", bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    body_font = Font(name="Times New Roman", size=11)
    body_align = Alignment(vertical="top", wrap_text=True)

    for row_idx, (doc_id, images, latex, created, updated) in enumerate(rows, 2):
        filename = Path(images).name if images and not images.startswith("http") else images
        values = [doc_id, filename, latex or "(chưa convert)", str(created), str(updated)]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = body_font
            cell.alignment = body_align
            cell.border = thin_border

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 80
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 20

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Đã export {len(rows)} bản ghi → {output_path}")


def main():
    cmd = sys.argv[1] if len(sys.argv) > 1 else ""

    if cmd == "insert":
        sources = sys.argv[2:]
        if not sources:
            print("Sử dụng: python convert.py insert <path_or_url> ...")
            sys.exit(1)
        app = App()
        try:
            app.insert_images(sources)
        finally:
            app.close()

    elif cmd == "export":
        out = sys.argv[2] if len(sys.argv) > 2 else "document_database.xlsx"
        export_to_excel(out)

    else:
        app = App()
        try:
            app.convert_all()
        finally:
            app.close()


if __name__ == "__main__":
    main()
